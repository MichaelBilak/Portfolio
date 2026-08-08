"""Generate DormUp Studio pricing workbook from data/pricing.ts values."""

from __future__ import annotations

import json
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    raise SystemExit("Run: pip install openpyxl")

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "data" / "DormUp-Studio-Pricing.xlsx"

# Mirrors data/pricing.ts + RU labels from translations
MAIN_SERVICES = [
    ("premium-site", "Custom Website Dev", 1299, "разово"),
    ("redesign", "Редизайн и оптимизация", 1499, "разово"),
    ("booking-flow", "Booking и lead-flow", 599, "разово"),
    ("monthly-support", "Ежемесячная поддержка", 199, "€/мес"),
    ("photo-video", "Фото и видео", 499, "разово"),
]

PACKAGES = [
    ("Custom Website Dev", "Starter", "starter", 1299, "разово", "нет", "Лендинг · 3–5 экранов · 1 язык · базовый SEO"),
    ("Custom Website Dev", "Business", "business", 1999, "разово", "да", "6–10 страниц · CMS · формы · SEO · SMB / hospitality"),
    ("Custom Website Dev", "Premium", "premium", 2999, "разово", "нет", "Полный сайт · мультиязычность · booking · анимации"),
    ("Редизайн и оптимизация", "Только аудит", "audit", 499, "разово", "нет", "Анализ · отчёт · roadmap"),
    ("Редизайн и оптимизация", "Standard", "standard", 1499, "разово", "да", "Редизайн ключевых экранов + внедрение"),
    ("Редизайн и оптимизация", "Полный", "full", 2499, "разово", "нет", "Полный редизайн · миграция · аналитика"),
    ("Booking и lead-flow", "Один поток", "single", 599, "разово", "нет", "Бронь или заявка · форма + CTA"),
    ("Booking и lead-flow", "Несколько потоков", "multi", 999, "разово", "да", "Интеграция booking · оптимизация CTA"),
    ("Booking и lead-flow", "Полный", "full", 1499, "разово", "нет", "Весь funnel · A/B · ежемесячный отчёт"),
    ("Ежемесячная поддержка", "Essential", "essential", 199, "€/мес", "нет", "~2 ч/мес · обновления · мониторинг"),
    ("Ежемесячная поддержка", "Growth", "growth", 399, "€/мес", "да", "~5 ч/мес · оптимизация · отчёт"),
    ("Ежемесячная поддержка", "Priority", "priority", 699, "€/мес", "нет", "~10 ч/мес · максимальный приоритет"),
    ("Фото и видео", "Полдня", "half-day", 499, "разово", "нет", "До 4 ч · фото или видео · базовый монтаж"),
    ("Фото и видео", "Полный день", "full-day", 899, "разово", "да", "Целый день · фото + видео"),
    ("Фото и видео", "Ретейнер", "retainer", 1499, "€/мес", "нет", "Регулярные съёмки · календарь контента"),
]

ADDONS = [
    ("Сайты", "Корпоративные", "corporate", "от", 1999, "разово"),
    ("Сайты", "Промо", "promo", "от", 999, "разово"),
    ("Сайты", "Лендинги", "landing", "от", 1299, "разово"),
    ("Сайты", "Медиа и блоги", "media-blog", "от", 2499, "разово"),
    ("Сайты", "No/Low-code", "no-code", "от", 899, "разово"),
    ("Продукты", "Веб-сервисы", "web-service", "от", 2999, "разово"),
    ("Продукты", "Интернет-магазины", "ecommerce", "от", 2499, "разово"),
    ("Продукты", "Личные кабинеты", "client-portal", "от", 1499, "разово"),
    ("Продукты", "Чат-боты", "chatbot", "от", 799, "разово"),
    ("Продукты", "Интранеты", "intranet", "от", 3499, "разово"),
    ("Продукты", "Мобильные приложения", "mobile-app", "от", 4999, "разово"),
    ("Дизайн", "UX & UI", "ux-ui", "от", 799, "разово"),
    ("Дизайн", "Брендинг", "branding", "от", 599, "разово"),
    ("Дизайн", "Motion & Sound", "motion-sound", "+", 399, "разово"),
    ("Дизайн", "UX-исследования", "ux-research", "от", 799, "разово"),
    ("Разработка", "CMS", "cms", "+", 299, "разово"),
    ("Разработка", "Мультиязычность", "multilingual", "+", 399, "за язык"),
    ("Разработка", "Backend / API", "backend", "+", 1499, "разово"),
    ("Разработка", "Quality Assurance", "qa", "+", 499, "разово"),
    ("Разработка", "DevOps", "devops", "+", 399, "разово"),
    ("Разработка", "Расширенный SEO", "seo-extended", "+", 299, "разово"),
]

HEADER_FILL = PatternFill("solid", fgColor="1A1F26")
HEADER_FONT = Font(bold=True, color="FCD34D", name="Calibri", size=11)
BODY_FONT = Font(name="Calibri", size=11)
MONEY_FMT = '#,##0 "€"'


def style_sheet(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def write_table(ws, title: str, headers: list[str], rows: list[tuple], widths: list[int], money_cols: set[int] | None = None) -> None:
    ws.append([title])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws["A1"].font = Font(bold=True, size=14, name="Calibri", color="F6F5F1")
    ws["A1"].fill = PatternFill("solid", fgColor="0F1620")
    ws.append(headers)
    for cell in ws[2]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    money_cols = money_cols or set()
    for row in rows:
        ws.append(list(row))
        r = ws.max_row
        for c in range(1, len(headers) + 1):
            cell = ws.cell(r, c)
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if c in money_cols:
                cell.number_format = MONEY_FMT
    style_sheet(ws, widths)


def main() -> None:
    wb = Workbook()
    wb.remove(wb.active)

    ws1 = wb.create_sheet("Основные услуги")
    write_table(
        ws1,
        "DormUp Studio — основные услуги (якорные цены «от»)",
        ["ID", "Услуга", "Цена от", "Периодичность", "Примечание"],
        [
            (sid, name, price, period, "Итоговая смета после брифа")
            for sid, name, price, period in MAIN_SERVICES
        ],
        [22, 34, 12, 14, 36],
        money_cols={3},
    )

    ws2 = wb.create_sheet("Пакеты")
    write_table(
        ws2,
        "Пакеты внутри основных услуг (не на сайте — для внутреннего прайса)",
        ["Услуга", "Пакет", "Tier ID", "Цена от", "Периодичность", "Популярный", "Что входит"],
        PACKAGES,
        [28, 18, 14, 12, 14, 12, 52],
        money_cols={4},
    )

    ws3 = wb.create_sheet("Доп. модули")
    write_table(
        ws3,
        "Дополнительные модули (надбавки к проекту)",
        ["Категория", "Модуль", "ID", "Тип", "Цена", "Периодичность", "Примечание"],
        [
            (*row[:5], row[5], "«от» — отдельный проект; «+» — к основной услуге")
            for row in ADDONS
        ],
        [16, 28, 18, 8, 12, 14, 40],
        money_cols={5},
    )

    ws4 = wb.create_sheet("Сводка")
    write_table(
        ws4,
        "Быстрая сводка",
        ["Тип", "Кол-во позиций", "Мин. цена", "Макс. цена"],
        [
            ("Основные услуги", len(MAIN_SERVICES), min(r[2] for r in MAIN_SERVICES), max(r[2] for r in MAIN_SERVICES)),
            ("Пакеты", len(PACKAGES), min(r[3] for r in PACKAGES), max(r[3] for r in PACKAGES)),
            ("Доп. модули", len(ADDONS), min(r[4] for r in ADDONS), max(r[4] for r in ADDONS)),
        ],
        [22, 16, 14, 14],
        money_cols={3, 4},
    )

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(json.dumps({"path": str(OUT)}))


if __name__ == "__main__":
    main()
